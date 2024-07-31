import os
import openpyxl
from copy import copy
import win32com.client as win32
from openpyxl import load_workbook
import sys

# Function to read employee details from Excel sheet
def read_Payroll(file_path):
    password = input("Enter password for the Excel file: ")
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False  # Keep Excel invisible
    wb = excel.Workbooks.Open(file_path, False, True, None, password)  
    ws_employee = wb["Payroll"]  # Access the sheet containing employee details
    ws_pdf = wb["Pay Slip"]      # Access the sheet containing PDF template
    details = []
    for row in ws_employee.iter_rows(min_row=7, values_only=True):
        details.append({'dep': row[1],'id': row[2], 'name': row[3], 'designation': row[4], 'salary': row[7], 'no_pay': row[13], 'br_allowance1': row[8], 'sa_ad': row[33], 'br_allowance2': row[9], 'paye': row[36], 'll': row[16], 'tra': row[65], 'tc': row[38], 'pa': row[66], 'sl': row[35], 'tfa': row[67], 'eale': row[29], 'va': row[24], 'bank': row[55], 'branch': row[57]})
    return details, ws_pdf, excel, wb

# Function to generate Excel files for each employee
def generate_excel(employee, template_sheet, output_folder):
    # Check if Employee ID is empty
    if not employee['id']:
        print(f"Skipping employee with empty ID: {employee['name']}")
        return

    # Create the output folder if it doesn't exist
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Create a copy of the template Excel file
    excel_name = f"{employee['id']}.xlsx"
    excel_path = os.path.join(output_folder, excel_name)

    # Copy the template sheet to a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in template_sheet.iter_rows(values_only=True):
        ws.append(row)
    
    # Copy cell styles from the template sheet to the new workbook
    for row in template_sheet.iter_rows(min_row=1, max_row=template_sheet.max_row, min_col=1, max_col=template_sheet.max_column):
        for cell in row:
            template_cell = template_sheet[cell.coordinate]
            new_cell = ws[cell.coordinate]
            new_cell.font = copy(template_cell.font)
            new_cell.fill = copy(template_cell.fill)
            new_cell.border = copy(template_cell.border)
            new_cell.alignment = copy(template_cell.alignment)
            new_cell.number_format = copy(template_cell.number_format)
            new_cell.protection = copy(template_cell.protection)
            
            # Copy column width
            column_letter = openpyxl.utils.get_column_letter(cell.column)
            ws.column_dimensions[column_letter].width = template_sheet.column_dimensions[column_letter].width

            # Copy merged cells
            if template_sheet.merged_cells.ranges:
                for merged_range in template_sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        ws.merge_cells(merged_range.coord)

            # Copy center alignment
            if template_cell.alignment.horizontal == 'center' or template_cell.alignment.vertical == 'center':
                ws[cell.coordinate].alignment = copy(template_cell.alignment)
    
    # Map data from employee details to the Excel template
    ws['B7'] = employee['dep']      # Department
    ws['B5'] = employee['id']       # Employee ID
    ws['B4'] = employee['name']     # Employee name
    ws['B6'] = employee['designation']    # Designation
    ws['B9'] = employee['salary']   # Basic salary
    ws['D9'] = employee['no_pay']   # No Pay Amount
    ws['B10'] = employee['br_allowance1']  # Budgetary Allowance 1
    ws['D10'] = employee['sa_ad']   # Salary Advance
    ws['B11'] = employee['br_allowance2']  # Budgetary Allowance 2
    ws['D11'] = employee['paye']    # Paye
    ws['B12'] = employee['ll']      # Lieu Leave
    ws['B13'] = employee['tra']     # Transport Allowance
    ws['D13'] = employee['tc']      # Telephone charges
    ws['B14'] = employee['pa']      # Performance Allowance 
    ws['D14'] = employee['sl']      # Staff Loan
    ws['B15'] = employee['tfa']     # Team Performance Allowance
    ws['B17'] = employee['eale']    # Earned Annual Leave Encashment
    ws['B18'] = employee['va']      # Vehicle Allowance
    ws['B28'] = employee['bank']    # Bank
    ws['C28'] = employee['branch']  # Branch

    # Save the Excel file
    wb.save(excel_path)

    return excel_path



# Main function to orchestrate the process
def main():
    excel_file = "PAYROLL FILE.xlsm"   # Specify the path to your Excel file containing both sheets
    output_folder = "output"            # Specify the path to the folder where you want to save the generated files
    
    result = read_Payroll(excel_file)
    if result is None:
        print("Error reading payroll.")
        return
    
    employees, template_sheet, excel, wb = result
    
    for employee in employees:
        generate_excel(employee, template_sheet, output_folder)
        print(f"Excel generated for Employee ID: {employee['id']}")
    
    # Close the workbook and quit Excel application properly
    if wb is not None:
        wb.Close(False)
    if excel is not None:
        excel.Quit()

if __name__ == "__main__":
    main()




# Function to convert Excel files to PDF and delete Excel files
def excel_to_pdf(input_folder, output_folder):
    excel_files = [f for f in os.listdir(input_folder) if f.endswith('.xlsx')]
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = False  # Keep Excel invisible

    for file in excel_files:
        excel_file = os.path.join(input_folder, file)
        pdf_file = os.path.join(output_folder, file.replace('.xlsx', '.pdf'))

        # Check if the Excel file exists
        if not os.path.exists(excel_file):
            print(f"Excel file '{excel_file}' not found. Skipping.")
            continue

        # Check if the output folder exists, create if not
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # Open Excel file
        try:
            wb = excel.Workbooks.Open(excel_file)
        except Exception as e:
            print(f"Error opening Excel file '{excel_file}': {e}")
            continue

        # Save as PDF
        try:
            wb.SaveAs(pdf_file, FileFormat=57)  # 57 is the PDF format code
        except Exception as e:
            print(f"Error saving PDF file for '{excel_file}': {e}")
            wb.Close()
            continue

        # Close Excel file
        wb.Close()

        # Delete Excel file
        try:
            os.remove(excel_file)
        except Exception as e:
            print(f"Error deleting Excel file '{excel_file}': {e}")

    excel.Quit()

# Get the directory of the executable file
exe_dir = os.path.dirname(sys.argv[0])

# Specify input and output folders using the executable file directory
input_folder = os.path.join(exe_dir, "output")
output_folder = os.path.join(exe_dir, "output")

# Convert Excel files to PDF and delete Excel files
excel_to_pdf(input_folder, output_folder)

print("Process Completed")